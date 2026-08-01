#!/usr/bin/env python3
"""Ежечасная синхронизация занятости из Google-таблиц отельеров в нашу шахматку.

Два формата таблиц-шахматок:
  • 'merge' — бронь = объединённая ячейка (Санрайз): дни в шапке, номер в столбце A.
  • 'color' — бронь = ЦВЕТ фона ячеек (Фемели): каждый номер = блок из N строк,
    занятый день = день-колонка с непустой заливкой.
Пишем занятость метками external_source='googlesheet_<tag>' (наши брони НЕ трогаем,
пересечения отсекаем — их отобьёт триггер А1). Идемпотентно: delete+insert.

Креды из окружения (GitHub secrets):
  GOOGLE_SA_B64 (base64 JSON) либо GOOGLE_SERVICE_ACCOUNT_JSON
  NEXT_PUBLIC_SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY
"""
import os, json, re, datetime, urllib.request, urllib.error, urllib.parse, base64
import gspread
from google.oauth2.service_account import Credentials
from google.auth.transport.requests import Request

YEAR = 2026
NIGHT = 86400

SOURCES = [
    {
        'name': 'Санрайз', 'tag': 'googlesheet_sunrise', 'mode': 'merge',
        'sheet_id': '16jmZEO_nWlZSY5hVS6F7rSzAxW9CRWlhppcV3XU-lms',
        'hotel_like': '%санрайз%', 'header_row': 0,
        'months': {'Май': 5, 'Июнь': 6, 'Июль': 7, 'Август': 8, 'Сентябрь': 9},
        'map': {'1': 1, '2': 2, '21': 3, '22': 4, '23': 5, '24': 6, '25': 7, '26': 8,
                '31': 9, '32': 10, '33': 11, '34': 12, '35': 13, '36': 14},
        'room_regex': r'номер\s*(\d+)', 'guest': 'Занято (Санрайз)',
    },
    {
        # Фемели: цветовой формат. Пока только ДОМИК 1-12 (люксы в таблице сведены
        # в 2 блока, у нас 8 номеров — сопоставление не решено, не трогаем).
        'name': 'Фемели', 'tag': 'googlesheet_femeli', 'mode': 'color',
        'sheet_id': '1q81E0jCexPCLelZJRJKF6TxYq_hXNPToIwk9701PFK0',
        'hotel_like': 'фемели', 'header_row': None,  # шапка = первая строка блока
        'months': {'МАЙ': 5, 'ИЮНЬ': 6, 'ИЮЛЬ': 7, 'АВГУСТ': 8, 'СЕНТЯБРЬ': 9,
                   'ОКТЯБРЬ': 10, 'НОЯБРЬ': 11},
        'label_prefix': 'ДОМИК', 'room_regex': r'Домик\s*(\d+)', 'guest': 'Занято (Фемели)',
    },
    {
        # Вилла Леона: шахматка в WPS 365 (не Google). Отельер включил
        # «Разрешить скачивание» — качаем xlsx по публичной ссылке и парсим.
        # Формат: столбец A=тип, B=№ номера, дни в строке 2, бронь = текст в
        # день заезда + ЗАЛИВКА до дня перед выездом (верим заливке; текстовую
        # запись «N-M» дотягиваем, если ячейки не закрашены). Лист = месяц,
        # с 2027 — суффикс «27».
        'name': 'Вилла Леона', 'tag': 'wps_villa_leona', 'mode': 'wps',
        'wps_sid': 'cbDaenFcrtTR3pSt',
        'hotel_like': 'вилла леона',
        'months': {'май': (2026, 5), 'июнь': (2026, 6), 'июль': (2026, 7), 'август': (2026, 8),
                   'сентябрь': (2026, 9), 'октябрь': (2026, 10), 'ноябрь': (2026, 11), 'декабрь': (2026, 12),
                   'январь27': (2027, 1), 'февраль27': (2027, 2), 'март27': (2027, 3), 'апрель27': (2027, 4),
                   'май27': (2027, 5), 'июнь27': (2027, 6), 'июль27': (2027, 7), 'август27': (2027, 8),
                   'сентябрь27': (2027, 9), 'октябрь27': (2027, 10), 'ноябрь27': (2027, 11), 'декабрь27': (2027, 12)},
        'room_regex': r'№(\d+)', 'guest': 'Занято (Вилла Леона)',
    },
]

SB_URL = os.environ['NEXT_PUBLIC_SUPABASE_URL'].rstrip('/')
SB_KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']

def sb(method, path, body=None, prefer=None):
    headers = {'apikey': SB_KEY, 'Authorization': 'Bearer ' + SB_KEY, 'Content-Type': 'application/json'}
    if prefer:
        headers['Prefer'] = prefer
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(SB_URL + path, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            txt = r.read().decode()
            return r.status, (json.loads(txt) if txt else None)
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()

def load_key():
    if os.environ.get('GOOGLE_SA_B64'):
        return json.loads(base64.b64decode(os.environ['GOOGLE_SA_B64']))
    return json.loads(os.environ['GOOGLE_SERVICE_ACCOUNT_JSON'])

def rooms_of(hotel_like, room_regex):
    st, hs = sb('GET', f"/rest/v1/hotels?select=id&title=ilike.{urllib.parse.quote(hotel_like)}")
    if not hs:
        raise SystemExit(f"отель не найден: {hotel_like}")
    hid = hs[0]['id']
    st, rs = sb('GET', f"/rest/v1/rooms?select=id,title,is_service&hotel_id=eq.{hid}")
    num2id = {}
    for r in rs or []:
        if r.get('is_service'):
            continue
        m = re.search(room_regex, r['title'] or '', re.IGNORECASE)
        if m:
            num2id[int(m.group(1))] = r['id']
    return num2id

def ci_unix(d): return int(datetime.datetime(d.year, d.month, d.day, 11, tzinfo=datetime.timezone.utc).timestamp())
def co_unix(d): return int(datetime.datetime(d.year, d.month, d.day, 9, tzinfo=datetime.timezone.utc).timestamp())

# ---------- MERGE (Санрайз): бронь = объединённая ячейка ----------
def parse_merge(gc, src):
    sh = gc.open_by_key(src['sheet_id'])
    meta = sh.fetch_sheet_metadata()
    def merges_of(t):
        sm = next((s for s in meta['sheets'] if s['properties']['title'] == t), None)
        return sm.get('merges', []) if sm else []
    hr = src['header_row']
    stays = []
    for title, mon in src['months'].items():
        try:
            ws = sh.worksheet(title)
        except gspread.WorksheetNotFound:
            continue
        vals = ws.get_all_values()
        if len(vals) <= hr:
            continue
        col2date = {}; prev = 0
        for c, v in enumerate(vals[hr]):
            v = v.strip()
            if not v.isdigit():
                continue
            d = int(v)
            if d < prev:
                break
            try:
                col2date[c] = datetime.date(YEAR, mon, d)
            except ValueError:
                break
            prev = d
        row2num = {r: src['map'][vals[r][0].strip()]
                   for r in range(hr + 1, len(vals)) if vals[r][0].strip() in src['map']}
        for m in merges_of(title):
            sr = m.get('startRowIndex', 0)
            if sr <= hr:
                continue
            er = m.get('endRowIndex', sr + 1); sc = m['startColumnIndex']; ec = m['endColumnIndex']
            num = next((row2num[rr] for rr in range(sr, er) if rr in row2num), None)
            if num is None:
                continue
            dates = sorted(col2date[c] for c in range(sc, ec) if c in col2date)
            if not dates:
                continue
            stays.append((num, dates[0], dates[-1] + datetime.timedelta(days=1)))
    return stays

# ---------- COLOR (Фемели): бронь = заливка фона ----------
def _is_white(bg):
    if not bg:
        return True
    return bg.get('red', 1) > 0.93 and bg.get('green', 1) > 0.93 and bg.get('blue', 1) > 0.93

def parse_color(token, src):
    occ = {}  # room number -> set(date)
    prefix = src['label_prefix']
    fields = 'sheets(data(rowData(values(formattedValue,effectiveFormat(backgroundColor)))))'
    for month, mon in src['months'].items():
        url = (f"https://sheets.googleapis.com/v4/spreadsheets/{src['sheet_id']}"
               f"?includeGridData=true&ranges={urllib.parse.quote(month)}!A1:AF160&fields={fields}")
        req = urllib.request.Request(url, headers={'Authorization': 'Bearer ' + token})
        try:
            with urllib.request.urlopen(req, timeout=60) as r:
                data = json.loads(r.read().decode())['sheets'][0]['data'][0].get('rowData', [])
        except (urllib.error.HTTPError, KeyError, IndexError):
            continue
        blocks = []
        for ri, row in enumerate(data):
            vals = row.get('values', [])
            c0 = (vals[0].get('formattedValue') if vals else '') or ''
            if c0.strip().upper().startswith(prefix):
                blocks.append((ri, c0.strip()))
        for bi, (sr, label) in enumerate(blocks):
            m = re.match(r'.*?(\d+)', label.upper())
            if not m:
                continue
            num = int(m.group(1))
            er = blocks[bi + 1][0] if bi + 1 < len(blocks) else sr + 5
            header = data[sr].get('values', [])
            col2day = {}; prev = 0
            for c in range(1, len(header)):
                v = (header[c].get('formattedValue') or '').strip()
                if not v.isdigit():
                    continue
                d = int(v)
                if d < prev:
                    break
                try:
                    datetime.date(YEAR, mon, d)
                except ValueError:
                    break
                col2day[c] = d; prev = d
            for r in range(sr, er):
                vv = data[r].get('values', []) if r < len(data) else []
                for c, day in col2day.items():
                    if c < len(vv) and not _is_white(vv[c].get('effectiveFormat', {}).get('backgroundColor')):
                        occ.setdefault(num, set()).add(datetime.date(YEAR, mon, day))
    # коалесценция занятых дат в интервалы
    stays = []
    for num, dates in occ.items():
        ds = sorted(dates); i = 0
        while i < len(ds):
            j = i
            while j + 1 < len(ds) and (ds[j + 1] - ds[j]).days == 1:
                j += 1
            stays.append((num, ds[i], ds[j] + datetime.timedelta(days=1))); i = j + 1
    return stays

# ---------- WPS (Вилла Леона): скачиваем xlsx по публичной ссылке ----------
def parse_wps(src):
    import io, tempfile
    import openpyxl
    ua = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/126.0 Safari/537.36'
    req = urllib.request.Request(
        f"https://eu.wps.com/office/docs/api/v3/office/file/{src['wps_sid']}/download",
        headers={'User-Agent': ua})
    meta = json.loads(urllib.request.urlopen(req, timeout=60).read().decode())
    url = meta.get('download_url') or meta.get('url')
    if not url:
        raise SystemExit(f"WPS не отдал ссылку на файл: {meta}")
    req2 = urllib.request.Request(url, headers={'User-Agent': ua})
    blob = urllib.request.urlopen(req2, timeout=120).read()
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)

    def is_filled(cell):
        f = cell.fill
        if not f or not f.patternType:
            return False
        v = f.fgColor.rgb
        return isinstance(v, str) and v not in ('00000000', 'FFFFFFFF')

    today = datetime.date.today()
    occ = {}
    for sheet, (Y, M) in src['months'].items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        col2day = {}
        prev = 0
        for c in range(3, ws.max_column + 1):
            v = ws.cell(2, c).value
            if not isinstance(v, (int, float)):
                break
            d = int(v)
            if d < prev:
                break
            try:
                datetime.date(Y, M, d)
            except ValueError:
                break
            col2day[c] = d
            prev = d
        for r in range(3, ws.max_row + 1):
            n = ws.cell(r, 2).value
            if not isinstance(n, (int, float)):
                continue
            n = int(n)
            days = occ.setdefault(n, set())
            for c, d in col2day.items():
                cell = ws.cell(r, c)
                txt = cell.value
                has = bool(txt is not None and str(txt).strip())
                if is_filled(cell) or has:
                    days.add(datetime.date(Y, M, d))
                if has:
                    m = re.search(r'(?<!\d)(\d{1,2})\s*[-–]\s*(\d{1,2})(?!\.\d)', str(txt))
                    if m:
                        a, b = int(m.group(1)), int(m.group(2))
                        if a == d and b > a:
                            for dd in range(a, b):
                                try:
                                    days.add(datetime.date(Y, M, dd))
                                except ValueError:
                                    pass

    stays = []
    for n, ds in occ.items():
        ds = sorted(x for x in ds if x >= today)
        i = 0
        while i < len(ds):
            j = i
            while j + 1 < len(ds) and (ds[j + 1] - ds[j]).days == 1:
                j += 1
            stays.append((n, ds[i], ds[j] + datetime.timedelta(days=1)))
            i = j + 1
    return stays

def sync_source(gc, token, src):
    num2id = rooms_of(src['hotel_like'], src['room_regex'])
    ids = list(num2id.values())
    if src['mode'] == 'merge':
        stays = parse_merge(gc, src)
    elif src['mode'] == 'color':
        stays = parse_color(token, src)
    else:
        stays = parse_wps(src)

    inlist = ",".join(f'"{i}"' for i in ids)
    st, rows = sb('GET', f"/rest/v1/reserves?select=room_id,start,end,external_source&room_id=in.({inlist})")
    ours = {i: set() for i in ids}
    for z in rows or []:
        if z.get('external_source') == src['tag']:
            continue
        for n in range(z['start'] // NIGHT, z['end'] // NIGHT):
            ours.setdefault(z['room_id'], set()).add(n)

    sb('DELETE', f"/rest/v1/reserves?external_source=eq.{src['tag']}&room_id=in.({inlist})", prefer='return=minimal')

    now = datetime.datetime.now(datetime.timezone.utc).isoformat()
    to_insert = []; skipped = 0
    for num, ci, co in stays:
        rid = num2id.get(num)
        if not rid:
            continue
        s = ci_unix(ci); e = co_unix(co)
        nights = set(range(s // NIGHT, e // NIGHT))
        if nights & ours.get(rid, set()):
            skipped += 1; continue
        to_insert.append({'room_id': rid, 'start': s, 'end': e, 'guest': src['guest'],
            'phone': '', 'price': 0, 'quantity': 1,
            'comment': f"Занятость из таблицы {src['name']} (авто, крон)",
            'created_by': src['tag'], 'edited_at': now, 'edited_by': src['tag'],
            'external_source': src['tag'], 'external_uid': f"{src['tag']}:{rid}:{s}-{e}",
            'external_feed_url': (
                'https://eu.wps.com/cms/module/common/preview/?sid=' + src['wps_sid']
                if src['mode'] == 'wps'
                else 'https://docs.google.com/spreadsheets/d/' + src['sheet_id']
            ),
            'external_synced_at': now})

    inserted = 0
    if to_insert:
        st, resp = sb('POST', "/rest/v1/reserves", body=to_insert, prefer='return=minimal')
        if st in (200, 201, 204):
            inserted = len(to_insert)
        else:
            for row in to_insert:
                st2, r2 = sb('POST', "/rest/v1/reserves", body=row, prefer='return=minimal')
                if st2 in (200, 201, 204):
                    inserted += 1
                elif isinstance(r2, str) and ('23P01' in r2 or 'Наложени' in r2):
                    skipped += 1
    return {'hotel': src['name'], 'parsed': len(stays), 'inserted': inserted, 'skipped': skipped}

def main():
    key = load_key()
    creds = Credentials.from_service_account_info(
        key, scopes=['https://www.googleapis.com/auth/spreadsheets.readonly'])
    gc = gspread.authorize(creds)
    creds.refresh(Request())
    token = creds.token
    summary = [sync_source(gc, token, src) for src in SOURCES]
    print(json.dumps({'status': 'ok', 'summary': summary}, ensure_ascii=False, indent=2))

if __name__ == '__main__':
    main()
